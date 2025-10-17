print("✅ Бот запущен на Render")
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters, CallbackQueryHandler
)
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Состояния диалога
PHOTO, SIZE, FORMAT, DETAILS, CONFIRM = range(5)

# Замените на ваш Telegram ID
ADMIN_CHAT_ID = "439141567"

# Стартовая команда
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['name'] = update.effective_user.full_name
    context.user_data['username'] = update.effective_user.username
    context.user_data['date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    await update.message.reply_text("Здравствуйте! Пожалуйста, отправьте изображение для вышивки.")
    return PHOTO

# Получение изображения (фото или файл)
async def receive_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.photo:
        photo = update.message.photo[-1]
        context.user_data['photo_id'] = photo.file_id
    elif update.message.document and update.message.document.mime_type.startswith("image/"):
        context.user_data['photo_id'] = update.message.document.file_id
    else:
        await update.message.reply_text("Пожалуйста, отправьте изображение (фото или файл).")
        return PHOTO

    await update.message.reply_text("Укажите желаемые размеры (например: 10x15 см)")
    return SIZE

# Получение размеров
async def receive_size(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['size'] = update.message.text
    await update.message.reply_text("Какой формат файла вам нужен? (например: DST, PES)")
    return FORMAT

# Получение формата
async def receive_format(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['format'] = update.message.text
    await update.message.reply_text("Есть ли дополнительные пожелания?")
    return DETAILS

# Получение пожеланий и кнопки подтверждения
async def receive_details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['details'] = update.message.text

    keyboard = [
        [InlineKeyboardButton("✅ Подтвердить", callback_data="confirm"),
         InlineKeyboardButton("❌ Отменить", callback_data="cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    summary = (
        f"📌 Заказ:\n"
        f"👤 Имя: {context.user_data['name']}\n"
        f"📅 Дата: {context.user_data['date']}\n"
        f"📐 Размер: {context.user_data['size']}\n"
        f"📁 Формат: {context.user_data['format']}\n"
        f"📝 Пожелания: {context.user_data['details']}\n"
        f"Нажмите кнопку ниже для подтверждения."
    )
    await update.message.reply_text(summary, reply_markup=reply_markup)
    return CONFIRM

# Обработка подтверждения или отмены
async def confirm_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "confirm":
        save_order_to_excel(context.user_data)

        await context.bot.send_photo(
            chat_id=ADMIN_CHAT_ID,
            photo=context.user_data['photo_id'],
            caption=(
                f"📥 Новый заказ от @{context.user_data['username']}:\n"
                f"Имя: {context.user_data['name']}\n"
                f"Дата: {context.user_data['date']}\n"
                f"Размер: {context.user_data['size']}\n"
                f"Формат: {context.user_data['format']}\n"
                f"Пожелания: {context.user_data['details']}"
            )
        )

        await query.edit_message_text("✅ Ваш заказ подтверждён и отправлен на обработку.")
    else:
        await query.edit_message_text("❌ Заказ отменён. Вы можете начать заново с /start.")

    return ConversationHandler.END

# Сохранение заказа в Excel
def save_order_to_excel(data):
    file_name = "orders.xlsx"
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Имя", "Username", "Дата", "Размер", "Формат", "Пожелания", "ID изображения"])

    ws.append([
        data.get("name", ""),
        data.get("username", ""),
        data.get("date", ""),
        data.get("size", ""),
        data.get("format", ""),
        data.get("details", ""),
        data.get("photo_id", "")
    ])
    wb.save(file_name)

# Запуск бота
import os
app = ApplicationBuilder().token(os.getenv("BOT_TOKEN")).build()

conv_handler = ConversationHandler(
    entry_points=[CommandHandler("start", start)],
    states={
        PHOTO: [MessageHandler(filters.PHOTO | filters.Document.IMAGE, receive_photo)],
        SIZE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_size)],
        FORMAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_format)],
        DETAILS: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_details)],
        CONFIRM: [CallbackQueryHandler(confirm_order)],
    },
    fallbacks=[]
)

# Добавление обработчика диалога
app.add_handler(conv_handler)

# Определение URL webhook — обязательно с HTTPS
WEBHOOK_URL = "https://embgallery.com/"

# Запуск бота через webhook
app.run_webhook(
    listen="0.0.0.0",
    port=int(os.environ.get("PORT", 10000)),
    webhook_url=WEBHOOK_URL
)


